"""Заполнить столбец «Участок» на листе «Ген. план» номерами участков из
«Квартал 1» (полосы под Айвазовского и Луначарского). Столбец «Дом» не трогаем.

Запуск:
    python scripts/fill_genplan_plots.py "data/Новая таблица.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Side

from lib import quarter_plots, workbook_edit

GENPLAN = "Ген. план"
QUARTER = "Квартал 1"
COL_PLOT = 11   # K
COL_HOUSE = 12  # L
START_ROW = 11  # заголовки в строке 10


def main(path: str):
    thin = Side(style="thin", color="FFBFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    with workbook_edit(path) as wb:
        plots_by_street = quarter_plots(wb[QUARTER])
        plots = sorted({p for lst in plots_by_street.values() for p in lst})

        gp = wb[GENPLAN]
        # снять объединения в K/L ниже заголовка
        for mr in [m for m in list(gp.merged_cells.ranges)
                   if m.min_col in (COL_PLOT, COL_HOUSE) and m.min_row >= START_ROW]:
            gp.unmerge_cells(str(mr))
        # очистить старое
        for r in range(START_ROW, START_ROW + 400):
            gp.cell(r, COL_PLOT).value = None
            gp.cell(r, COL_HOUSE).value = None

        row = START_ROW
        for pid in plots:
            ck = gp.cell(row, COL_PLOT, pid)
            ck.border = border
            ck.alignment = center
            cl = gp.cell(row, COL_HOUSE)
            cl.border = border
            cl.alignment = center
            row += 1

    print(f"Записано участков: {len(plots)} ({START_ROW}..{row - 1})")
    for street, lst in plots_by_street.items():
        print(f"  {street}: {len(lst)}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else str(
        Path(__file__).resolve().parents[1] / "data" / "Новая таблица.xlsx"
    )
    main(path)
