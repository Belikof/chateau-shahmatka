"""Общие утилиты для работы с шахматкой Château Del Mare.

openpyxl при сохранении теряет картинки/чертежи (drawings). Поэтому весь
пайплайн: сделать бэкап -> править через openpyxl -> save -> вернуть media и
drawings из бэкапа через прямую пересборку zip (restore_drawings).
"""
from __future__ import annotations

import re
import shutil
import zipfile
from pathlib import Path

# Каталог домов: (Проект, этажей, дом м², черновая, ПЧО, под ключ)
CATALOG = [
    ("Костеро", 1, 32, 5379000, 5709000, 6094000),
    ("Линия 40", 1, 40, 7000000, 7600000, 8500000),
    ("Линия 72", 1, 72, 9190000, 9990000, 11260000),
    ("Логос", 1, 64, 8300000, 9090000, 10990000),
    ("Логос мини", 1, 40, 7000000, 7600000, 8500000),
    ("Палаццо", 2, 80, 9490000, 10190000, 11990000),
    ("Эдем", 2, 60, 8300000, 9090000, 10990000),
    ("Арбор", 1, 32, 5379000, 5819000, 6369000),
    ("Баня", 1, 24, 2500000, 2800000, 3200000),
    ("Вилла Маре", 2, 136, 10390000, 11100000, 12850000),
]

CORNER_RATE = 1000          # надбавка к дому за м² на угловом участке
BASE_LAND_M2 = 500          # участок, входящий в цену дома по каталогу
ACTIVE_STREETS = ("Айвазовского", "Луначарского")

DRAWING_TAG = (
    '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/'
    '2006/relationships" r:id="{rid}"/>'
)


def fmt(n) -> str:
    """100000 -> '100 000'."""
    if n is None:
        return "—"
    return f"{int(round(n)):,}".replace(",", " ")


def cell_rgb(cell):
    f = cell.fill
    if not f or f.fill_type != "solid":
        return None
    c = f.start_color
    return c.rgb if c.type == "rgb" else None


def quarter_plots(ws, lower="Дивноморская"):
    """Возвращает {улица: [номера участков]} для активных улиц на листе квартала.

    Берёт все числа > 1000 между меткой улицы и следующей меткой.
    """
    order = list(ACTIVE_STREETS) + [lower]
    label_row = {}
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and v.strip() in order:
                label_row.setdefault(v.strip(), r)

    result = {}
    for i, street in enumerate(ACTIVE_STREETS):
        r0 = label_row.get(street)
        r1 = label_row.get(order[i + 1])
        if r0 is None or r1 is None:
            result[street] = []
            continue
        plots = []
        for r in range(r0 + 1, r1):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and v > 1000:
                    plots.append(int(v))
        result[street] = sorted(plots)
    return result


def _norm(target: str) -> str:
    target = target.lstrip("/")
    return target if target.startswith("xl/") else "xl/" + target


def _sheet_paths(parts: dict) -> dict:
    wbxml = parts["xl/workbook.xml"].decode()
    rels = parts["xl/_rels/workbook.xml.rels"].decode()
    rid = {}
    for m in re.finditer(r"<Relationship\b[^>]*>", rels):
        tag = m.group(0)
        i = re.search(r'Id="(rId\d+)"', tag)
        t = re.search(r'Target="([^"]+)"', tag)
        if i and t:
            rid[i.group(1)] = _norm(t.group(1))
    out = {}
    for m in re.finditer(r"<sheet\b[^>]*>", wbxml):
        tag = m.group(0)
        nm = re.search(r'name="([^"]+)"', tag)
        ri = re.search(r'r:id="(rId\d+)"', tag)
        if nm and ri and ri.group(1) in rid:
            out[nm.group(1)] = rid[ri.group(1)]
    return out


def restore_drawings(backup: str | Path, target: str | Path):
    """Вернуть media и drawings из backup в target (после openpyxl save)."""
    backup, target = str(backup), str(target)
    with zipfile.ZipFile(backup) as zo, zipfile.ZipFile(target) as zm:
        o = {n: zo.read(n) for n in zo.namelist()}
        m = {n: zm.read(n) for n in zm.namelist()}

    for n, d in o.items():
        if (
            n.startswith("xl/media/")
            or re.match(r"xl/drawings/drawing\d+\.xml$", n)
            or re.match(r"xl/drawings/_rels/drawing\d+\.xml\.rels$", n)
        ):
            m[n] = d

    # карта лист -> drawing берётся из бэкапа (исходные связи)
    sp_target = _sheet_paths(m)
    sp_backup = _sheet_paths(o)
    sheet_drawing = {}
    for sname, spath in sp_backup.items():
        relf = spath.replace("worksheets/", "worksheets/_rels/") + ".rels"
        if relf in o:
            dm = re.search(r"drawings/(drawing\d+\.xml)", o[relf].decode())
            if dm:
                sheet_drawing[sname] = dm.group(1)

    for sname, dfile in sheet_drawing.items():
        if sname not in sp_target or ("xl/drawings/" + dfile) not in m:
            continue
        sx = sp_target[sname]
        rx = sx.replace("worksheets/", "worksheets/_rels/") + ".rels"
        xml = m[sx].decode()
        rels = m.get(
            rx,
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<Relationships xmlns="http://schemas.openxmlformats.org/'
            b'package/2006/relationships"></Relationships>',
        ).decode()
        if "relationships/drawing" not in rels:
            ids = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels)]
            rid = f"rId{max(ids, default=0) + 1}"
            rels = rels.replace(
                "</Relationships>",
                f'<Relationship Id="{rid}" Type="http://schemas.openxmlformats.org/'
                f'officeDocument/2006/relationships/drawing" '
                f'Target="../drawings/{dfile}"/></Relationships>',
            )
        else:
            mm = re.search(r'Id="(rId\d+)"[^>]*relationships/drawing', rels) or re.search(
                r'relationships/drawing"[^>]*Id="(rId\d+)"', rels
            )
            rid = mm.group(1)
        if "<drawing" not in xml:
            xml = xml.replace("</worksheet>", DRAWING_TAG.format(rid=rid) + "</worksheet>")
        xml = re.sub(r'<drawing r:id="(rId\d+)"/>', DRAWING_TAG.format(rid=r"\1"), xml)
        m[sx] = xml.encode()
        m[rx] = rels.encode()

    # content types: вернуть Override для drawings/media + Default png
    ct = m["[Content_Types].xml"].decode()
    oct = o["[Content_Types].xml"].decode()
    for part in re.findall(
        r'<Override[^>]+PartName="(/xl/(?:drawings/drawing\d+\.xml|media/image\d+\.\w+))"[^>]*/>',
        oct,
    ):
        block = re.search(
            r'<Override[^>]+PartName="' + re.escape(part) + r'"[^>]*/>', oct
        ).group(0)
        if block not in ct:
            ct = ct.replace("</Types>", block + "</Types>")
    if 'Extension="png"' not in ct:
        ct = ct.replace(
            "</Types>", '<Default Extension="png" ContentType="image/png"/></Types>'
        )
    m["[Content_Types].xml"] = ct.encode()

    with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as z:
        for n, d in m.items():
            z.writestr(n, d)


class workbook_edit:
    """Контекст: бэкап -> правки -> save -> restore_drawings -> удалить бэкап.

    Пример:
        with workbook_edit(path) as wb:
            ws = wb["Квартал 1"]
            ...
    """

    def __init__(self, path: str | Path):
        from openpyxl import load_workbook

        self.path = Path(path)
        self.backup = self.path.with_suffix(".bak.xlsx")
        shutil.copy2(self.path, self.backup)
        self.wb = load_workbook(self.path)

    def __enter__(self):
        return self.wb

    def __exit__(self, exc_type, exc, tb):
        if exc_type is None:
            self.wb.save(self.path)
            restore_drawings(self.backup, self.path)
        if self.backup.exists():
            self.backup.unlink()
        return False
