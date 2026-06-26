"""Калькулятор на листе «Квартал 1» (строки 29–30).

Поля (выпадающие списки):
    Участок  — только участки из «Квартал 1» (Айвазовского + Луначарского)
    Дом      — проекты из «Каталог»
    Отделка  — Черновая / ПЧО / Под ключ
    Способ покупки — Наличные / Ипотека / Рассрочка (пока не влияет на сумму)
    ИТОГ     — цена земли + угловая надбавка + цена дома по отделке

Список участков пишется на скрытый лист «Данные участки» (столбец P) и
подключается к выпадающему списку через именованный диапазон.

Запуск:
    python scripts/build_calculator.py "data/Новая таблица.xlsx"
"""
from __future__ import annotations

import sys
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from lib import quarter_plots, workbook_edit

QUARTER = "Квартал 1"
DATA = "Данные участки"
CALC_ROW_LABEL = 29
CALC_ROW_INPUT = 30
HELP_COL = 16  # P на «Данные участки»

RESULT_FORMULA = (
    '=IF(OR($B$30="",$C$30="",$D$30=""),"— выберите участок, дом и отделку —",'
    "IFERROR(VLOOKUP($B$30,'Данные участки'!$A:$L,11,0)*1,0)"
    "+IFERROR(VLOOKUP($B$30,'Данные участки'!$A:$L,12,0)*1,0)"
    "+IFERROR(INDEX(Каталог!$D:$F,MATCH($C$30,Каталог!$A:$A,0),"
    'IF($D$30="Черновая",1,IF($D$30="ПЧО",2,3)))*1,0))'
)


def _set_name(wb, name, ref):
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
    except Exception:
        pass
    try:
        wb.defined_names.add(DefinedName(name, attr_text=ref))
    except Exception:
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


def main(path: str):
    thin = Side(style="thin", color="FF808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lab_font = Font(bold=True, size=10, color="FFFFFFFF")
    lab_fill = PatternFill("solid", fgColor="FF4472C4")
    in_fill = PatternFill("solid", fgColor="FFFFF2CC")
    res_fill = PatternFill("solid", fgColor="FFD9EAD3")

    with workbook_edit(path) as wb:
        plots = sorted({p for lst in quarter_plots(wb[QUARTER]).values() for p in lst})

        # helper-список участков на скрытом листе
        du = wb[DATA]
        du.cell(1, HELP_COL, "Участки (калькулятор)")
        for r in range(2, 500):
            du.cell(r, HELP_COL).value = None
        for i, pid in enumerate(plots, 2):
            du.cell(i, HELP_COL, pid)
        last = len(plots) + 1

        _set_name(wb, "СписокУчастков", f"'Данные участки'!$P$2:$P${last}")
        _set_name(wb, "СписокДомов", "Каталог!$A$2:$A$11")

        kv = wb[QUARTER]
        kv.cell(CALC_ROW_LABEL, 1, "КАЛЬКУЛЯТОР").font = Font(bold=True, size=12)
        for col, txt in [(2, "Участок"), (3, "Дом"), (4, "Отделка"),
                         (5, "Способ покупки"), (6, "ИТОГ, ₽")]:
            c = kv.cell(CALC_ROW_LABEL, col, txt)
            c.font = lab_font
            c.fill = lab_fill
            c.alignment = center
            c.border = border
        for col in (2, 3, 4, 5):
            c = kv.cell(CALC_ROW_INPUT, col)
            c.fill = in_fill
            c.alignment = center
            c.border = border
        res = kv.cell(CALC_ROW_INPUT, 6)
        res.fill = res_fill
        res.alignment = center
        res.border = border
        res.font = Font(bold=True, size=11)
        res.number_format = '#,##0" ₽"'
        res.value = RESULT_FORMULA

        for col, w in [(2, 12), (3, 14), (4, 12), (5, 16), (6, 16)]:
            kv.column_dimensions[chr(64 + col)].width = w

        dvs = [
            (DataValidation(type="list", formula1="СписокУчастков", allow_blank=True), "B30"),
            (DataValidation(type="list", formula1="СписокДомов", allow_blank=True), "C30"),
            (DataValidation(type="list", formula1='"Черновая,ПЧО,Под ключ"', allow_blank=True), "D30"),
            (DataValidation(type="list", formula1='"Наличные,Ипотека,Рассрочка"', allow_blank=True), "E30"),
        ]
        for dv, addr in dvs:
            kv.add_data_validation(dv)
            dv.add(addr)

    print(f"Калькулятор готов. Участков в списке: {len(plots)}")


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else str(
        Path(__file__).resolve().parents[1] / "data" / "Новая таблица.xlsx"
    )
    main(path)
