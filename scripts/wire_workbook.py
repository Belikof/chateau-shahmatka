"""Связать листы: зоны по «Виду», цвета статусов на «Квартал 1», тексты для людей."""
from __future__ import annotations

import re

import scripts.gsheet as g
from scripts.style_genplan_tables import main as style_genplan

QUARTER = "Квартал 1"
DATA = "Данные участки"
GUIDE = "Инструкция для брокеров"

# цвета статусов (RGB 0–1)
STATUS_CLR = {
    "Продано": {"red": 0.88, "green": 0.4, "blue": 0.4},
    "Бронь": {"red": 1.0, "green": 0.85, "blue": 0.4},
    "Резерв": {"red": 0.96, "green": 0.7, "blue": 0.45},
}

INSTRUCTION = (
    "КАК ПОСТАВИТЬ БРОНЬ\n\n"
    "1. Найдите участок в списке ниже (можно отфильтровать по номеру или улице)\n"
    "2. В столбце «Статус» выберите «Бронь»\n"
    "3. Нажмите на номер участка → вставьте примечание: дата, брокер, срок брони\n"
    "4. Цвет на листе «Квартал 1» обновится сам\n\n"
    "Свободно — в продаже  |  Бронь — забронирован  |  "
    "Резерв — думает  |  Продано — закрыто"
)

GUIDE_LINES = [
    ["ИНСТРУКЦИЯ — Шато Дель Маре"],
    [""],
    ["Где смотреть лоты"],
    ["• «Квартал 1» — шахматка с номерами и площадями"],
    ["• «Ген. план» — карта посёлка"],
    ["• «Данные участки» — сюда вносим бронь и статусы"],
    ["• Материалы: https://belikof.github.io/chateau-shahmatka/"],
    [""],
    ["Как поставить бронь"],
    ["1. Откройте лист «Данные участки»."],
    ["2. В таблице найдите нужный номер (фильтр в шапке таблицы)."],
    ["3. В столбце «Статус» выберите «Бронь»."],
    ["4. К номеру участка добавьте примечание, например:"],
    ["   12.07, Иванова, брокер Смирнов, бронь до 19.07"],
    ["5. На листе «Квартал 1» ячейка станет жёлтой."],
    [""],
    ["Цвета на «Квартал 1»"],
    ["• Белый — свободен"],
    ["• Жёлтый — бронь"],
    ["• Оранжевый — резерв"],
    ["• Красный — продан"],
    [""],
    ["Статусы"],
    ["• Резерв — клиент думает"],
    ["• Свободно — снова в продаже"],
    ["• Продано — только куратор"],
]

QUARTER_LEGEND = [
    ["Цвета на этой шахматке (обновляются сами из «Данные участки»):"],
    ["Белый — свободен  |  Жёлтый — бронь  |  Оранжевый — резерв  |  Красный — продан"],
    ["Бронь и резерв — лист «Данные участки», столбец «Статус»"],
]


def _sid(ss, title: str) -> int:
    for s in ss.fetch_sheet_metadata()["sheets"]:
        if s["properties"]["title"] == title:
            return s["properties"]["sheetId"]
    raise KeyError(title)


def _plot_grid_range(ss) -> dict:
    """Диапазон ячеек с номерами участков на «Квартал 1»."""
    vals = g.worksheet(QUARTER).get_all_values()
    min_r, max_r, min_c, max_c = 999, 0, 999, 0
    for r, row in enumerate(vals, 1):
        for c, v in enumerate(row, 1):
            if v and re.match(r"^\d{4}", str(v).strip()):
                if r == 15:  # строка калькулятора
                    continue
                min_r, max_r = min(min_r, r), max(max_r, r)
                min_c, max_c = min(min_c, c), max(max_c, c)
    if min_r > max_r:
        min_r, max_r, min_c, max_c = 2, 14, 1, 57
    return {
        "sheetId": _sid(ss, QUARTER),
        "startRowIndex": min_r - 1,
        "endRowIndex": max_r,
        "startColumnIndex": min_c - 1,
        "endColumnIndex": max_c,
    }


HELPER_COL = 27  # AA
HELPER_START = 17
HELPER_END = 143  # 127 участков


def _data_bounds(ss) -> tuple[int, int]:
    """Строки данных на «Данные участки» (1-based)."""
    rows = g.worksheet(DATA).col_values(1)
    hdr = next((i + 1 for i, v in enumerate(rows) if v.strip() == "№ участка"), 8)
    n = sum(1 for v in rows[hdr:] if v and str(v).strip().isdigit())
    return hdr + 1, hdr + n


def setup_status_helper(ss) -> str:
    """Скрытая таблица №→статус на «Квартал 1»."""
    ws = g.worksheet(QUARTER)
    sid = _sid(ss, QUARTER)
    d0, d1 = _data_bounds(ss)
    n = d1 - d0 + 1
    end = HELPER_START + n - 1

    from openpyxl.utils import get_column_letter

    c = get_column_letter(HELPER_COL)
    c2 = get_column_letter(HELPER_COL + 1)
    ws.update(
        range_name=f"{c}{HELPER_START}:{c2}{end}",
        values=[
            [
                f"=INDEX('{DATA}'!$A${d0}:$A${d1};ROW()-{HELPER_START}+1)",
                f"=INDEX('{DATA}'!$E${d0}:$E${d1};ROW()-{HELPER_START}+1)",
            ]
            for _ in range(n)
        ],
        value_input_option="USER_ENTERED",
    )

    ss.batch_update(
        {
            "requests": [
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sid,
                            "dimension": "COLUMNS",
                            "startIndex": HELPER_COL - 1,
                            "endIndex": HELPER_COL + 1,
                        },
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                }
            ]
        }
    )
    return f"${c}${HELPER_START}:${c2}${end}"


def _status_formula(cell: str, status: str, lookup_range: str) -> str:
    return (
        f"=И(ЕЧИСЛО(ЗНАЧЕН(ЛЕВСИМВ({cell};4)));"
        f"ЕСЛИОШИБКА(ВПР(ЗНАЧЕН(ЛЕВСИМВ({cell};4));{lookup_range};2;0)=\"{status}\";ЛОЖЬ))"
    )


def apply_quarter_status_colors(ss) -> None:
    sid = _sid(ss, QUARTER)
    lookup = setup_status_helper(ss)
    grid = _plot_grid_range(ss)
    tl_row = grid["startRowIndex"] + 1
    tl_col = grid["startColumnIndex"]
    from openpyxl.utils import get_column_letter

    tl_a1 = f"{get_column_letter(tl_col + 1)}{tl_row}"

    meta = ss.fetch_sheet_metadata()
    n_rules = 0
    for s in meta["sheets"]:
        if s["properties"]["sheetId"] == sid:
            n_rules = len(s.get("conditionalFormats", []))
            break

    reqs = [
        {"deleteConditionalFormatRule": {"sheetId": sid, "index": i}}
        for i in range(n_rules - 1, -1, -1)
    ]

    for status in ("Продано", "Бронь", "Резерв"):
        reqs.append(
            {
                "addConditionalFormatRule": {
                    "rule": {
                        "ranges": [grid],
                        "booleanRule": {
                            "condition": {
                                "type": "CUSTOM_FORMULA",
                                "values": [
                                    {"userEnteredValue": _status_formula(tl_a1, status, lookup)}
                                ],
                            },
                            "format": {"backgroundColor": STATUS_CLR[status]},
                        },
                    },
                    "index": 0,
                }
            }
        )

    if reqs:
        ss.batch_update({"requests": reqs})


def hide_tech_columns(ss) -> None:
    sid = _sid(ss, DATA)
    ss.batch_update(
        {
            "requests": [
                {
                    "updateDimensionProperties": {
                        "range": {
                            "sheetId": sid,
                            "dimension": "COLUMNS",
                            "startIndex": 6,
                            "endIndex": 8,
                        },
                        "properties": {"hiddenByUser": True},
                        "fields": "hiddenByUser",
                    }
                },
            ]
        }
    )


def update_texts(ss) -> None:
    ws = g.worksheet(DATA)
    ws.update(range_name="A1", values=[[INSTRUCTION]], value_input_option="USER_ENTERED")

    guide = ss.worksheet(GUIDE)
    guide.clear()
    guide.update(range_name="A1", values=GUIDE_LINES, value_input_option="USER_ENTERED")

    q = g.worksheet(QUARTER)
    q.batch_update(
        [
            {"range": "A11", "values": [[QUARTER_LEGEND[0][0]]]},
            {"range": "A12", "values": [[QUARTER_LEGEND[1][0]]]},
            {"range": "A13", "values": [[QUARTER_LEGEND[2][0]]]},
            {"range": "A14", "values": [[""]]},
            {"range": "A15", "values": [[""]]},
        ],
        value_input_option="USER_ENTERED",
    )


def main() -> None:
    ss = g.open_spreadsheet()
    print("Тексты и легенды…")
    update_texts(ss)
    print("Скрыть технические столбцы G–H…")
    hide_tech_columns(ss)
    print("Генплан: зоны по «Виду»…")
    style_genplan()
    print("Цвета статусов на «Квартал 1»…")
    apply_quarter_status_colors(ss)
    print("Готово.")


if __name__ == "__main__":
    main()
